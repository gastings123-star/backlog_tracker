"""
Генерация шаблона импорта задач для трекера согласования бэклога.
Колонки и их порядок соответствуют автосопоставлению (populateColMap)
в index.html.

Запуск:  python3 scripts/build_template.py
Результат: templates/tasks-import-template.xlsx
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "templates", "tasks-import-template.xlsx")

HEADERS = [
    ("Название",             35, "Краткое название задачи. Обязательное поле."),
    ("Анализ (дни)",         14, "Оценка аналитики в днях. Целое или дробное число."),
    ("Разработка (дни)",     16, "Оценка разработки в днях."),
    ("Тестирование (дни)",   18, "Оценка тестирования в днях."),
    ("Приоритет",            14, "Высокий / Средний / Низкий (можно High/Medium/Low). Пусто = не задан."),
    ("Ценность для бизнеса", 32, "Какую пользу приносит задача."),
    ("Экон. эффект",         22, "Например: ~2 млн руб/год."),
    ("Смежные системы",      24, "Через запятую: CRM, АБС, SWIFT."),
    ("Стейкхолдер",          22, "ФИО ответственного со стороны бизнеса."),
    ("Jira",                 16, "Ключ задачи в Jira, напр. ABC-123."),
    ("SD",                   16, "Номер заявки Service Desk."),
    ("Комментарий",          30, "Любые дополнительные заметки."),
]

EXAMPLES = [
    ["Онбординг клиентов",      5,  15, 8,  "Высокий", "Сокращение времени онбординга на 40%", "~2 млн руб/год", "CRM, АБС",   "Иванов А.П.",  "ABC-101", "SD-5501", ""],
    ["Отчёт по транзакциям",    3,  10, 5,  "Средний", "Автоматизация отчётности",              "",                "АБС",         "Петрова М.В.", "ABC-102", "",        ""],
    ["API интеграция с АБС",    8,  30, 12, "Высокий", "Ускорение обработки платежей",          "~5 млн руб/год", "АБС, SWIFT", "Сидоров К.Л.", "ABC-103", "SD-5610", "Требует согласования с ИБ"],
    ["Личный кабинет 2.0",     12,  40, 18, "Низкий",  "Улучшение UX для клиентов",             "",                "CRM",         "Козлов Д.И.",  "",         "",        ""],
]

wb = Workbook()

# === Лист 1: Задачи =========================================================
ws = wb.active
ws.title = "Задачи"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="534AB7")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, (title, width, hint) in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    cell.comment = Comment(hint, "Шаблон")
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Пометить обязательную колонку (Название) красной точкой через формат шапки
ws.cell(row=1, column=1).value = "Название *"

ws.row_dimensions[1].height = 32

# Примеры
example_fill = PatternFill("solid", fgColor="F7F7FB")
for r_idx, row in enumerate(EXAMPLES, start=2):
    for c_idx, val in enumerate(row, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        c.fill = example_fill

# Числовой формат для колонок дней (B, C, D)
for col_letter in ("B", "C", "D"):
    for r in range(2, 2 + len(EXAMPLES) + 200):  # с запасом
        ws[f"{col_letter}{r}"].number_format = "0.##"

# Пустые строки для заполнения с границами
for r in range(2 + len(EXAMPLES), 2 + len(EXAMPLES) + 30):
    for c_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=r, column=c_idx).border = border

# === Лист 2: Справочники ====================================================
ref = wb.create_sheet("Справочники")
ref.sheet_view.showGridLines = False
ref["A1"] = "Приоритет"
ref["A1"].font = Font(bold=True)
priorities = ["Высокий", "Средний", "Низкий"]
for i, v in enumerate(priorities, start=2):
    ref.cell(row=i, column=1, value=v)
ref.column_dimensions["A"].width = 18

# Валидация значения «Приоритет» (колонка E на листе «Задачи»)
dv = DataValidation(
    type="list",
    formula1="=Справочники!$A$2:$A$4",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Недопустимое значение",
    error="Допустимо: Высокий, Средний, Низкий (или пусто).",
)
dv.add("E2:E1000")
ws.add_data_validation(dv)

# Валидация чисел для дней
dv_num = DataValidation(
    type="decimal",
    operator="greaterThanOrEqual",
    formula1=0,
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Нужно число",
    error="Введите неотрицательное число дней.",
)
dv_num.add("B2:D1000")
ws.add_data_validation(dv_num)

# === Лист 3: Инструкция =====================================================
info = wb.create_sheet("Инструкция", 0)
info.sheet_view.showGridLines = False
info.column_dimensions["A"].width = 110

lines = [
    ("Шаблон импорта задач — Трекер согласования бэклога", True),
    ("", False),
    ("1. Заполняйте задачи на листе «Задачи», начиная со строки 2.", False),
    ("2. Обязательная колонка — «Название». Строки без названия игнорируются.", False),
    ("3. Дни (анализ / разработка / тестирование) — числа, можно дробные (например, 2.5).", False),
    ("4. Приоритет — выберите из списка: Высокий / Средний / Низкий. Можно оставить пустым.", False),
    ("   Также распознаются англ. варианты: High / Medium / Low.", False),
    ("5. Поля «Jira», «SD», «Стейкхолдер», «Системы» и т.п. — свободный текст.", False),
    ("6. Лишние колонки можно оставить — при импорте сопоставление подхватится автоматически.", False),
    ("7. Сохраните файл и загрузите его в трекер: кнопка «Импорт» → выбрать .xlsx.", False),
    ("", False),
    ("Импортируемые задачи попадают в статус «На рассмотрении» без привязки к команде.", False),
    ("Назначить команду и утвердить можно уже внутри трекера.", False),
]
for i, (text, bold) in enumerate(lines, start=1):
    c = info.cell(row=i, column=1, value=text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if bold or i == 1:
        c.font = Font(bold=True, size=13 if i == 1 else 11)

wb.save(OUT)
print(f"OK -> {OUT}")
